"""Job UI routes and query helpers."""

import io
import json
from datetime import datetime
from urllib.parse import urlencode

import psycopg2
import psycopg2.extras
from fastapi import APIRouter, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ui import schema
from ui.config import EVENTS_TABLE, TABLE
from ui.db import execute, fetch_all, fetch_one, get_db
from ui.rendering import format_salary, render
from ui.services import stats as stats_service


router = APIRouter()

REVIEW_ACTIONS = {
    "apply_target": {
        "decision": "PASS",
        "user_status": "interested",
        "label": "Reviewed: apply target",
    },
    "maybe": {
        "decision": "MAYBE",
        "user_status": "interested",
        "label": "Reviewed: maybe",
    },
    "skip": {
        "decision": "SKIP",
        "user_status": "dismissed",
        "label": "Reviewed: skip",
    },
}


def _review_notes_value(existing_notes, submitted_notes):
    submitted = (submitted_notes or "").strip()
    if submitted:
        return submitted
    return existing_notes if existing_notes else None


def _query_bool(request: Request, name: str, default: bool = False) -> bool:
    value = request.query_params.get(name)
    if value is None:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def _jobs_filter_context(request: Request):
    filters = {
        "view": request.query_params.get("view", ""),
        "decision": request.query_params.get("decision", ""),
        "applied": request.query_params.get("applied", ""),
        "min_score": request.query_params.get("min_score", "0"),
        "max_score": request.query_params.get("max_score", "10"),
        "search": request.query_params.get("search", ""),
        "hide_staffing": _query_bool(request, "hide_staffing"),
        "hide_geo": _query_bool(request, "hide_geo"),
        "english_only": _query_bool(request, "english_only"),
        "recent_only": _query_bool(request, "recent_only", True),
        "hide_rejected": _query_bool(request, "hide_rejected", True),
        "sort": request.query_params.get("sort", "fit_score"),
        "order": request.query_params.get("order", "desc"),
    }
    query = urlencode({
        key: str(value).lower() if isinstance(value, bool) else value
        for key, value in filters.items()
        if value not in ("", None)
    })
    return {"filters": filters, "jobs_query_string": query}


def build_order_clause(sort_col: str, sort_dir: str, view: str = "") -> str:
    if view == "review":
        return "COALESCE(fit_score, 0) DESC, analyzed_at DESC NULLS LAST, id DESC"
    return f"{sort_col} {sort_dir}, id {sort_dir}"


def build_job_filter(decision="", applied="", min_score=0, max_score=10, search="", view="",
                     hide_staffing=False, hide_geo=False, english_only=False,
                     hide_rejected=False, recent_only=True):
    params: list = []
    if view == "error":
        conditions = ["status = 'error'"]
    elif view == "dead":
        conditions = ["status = 'dead'"]
    elif view == "failed":
        conditions = ["status IN ('error', 'dead')"]
    elif view == "duplicates":
        conditions = ["possible_duplicate_of IS NOT NULL AND duplicate_confirmed IS NULL"]
    elif view == "review":
        if schema.HAS_HUMAN_REVIEW_COLUMNS:
            conditions = ["status = 'analyzed' AND (needs_human_review = TRUE OR decision = 'pending_review')"]
        else:
            conditions = ["status = 'analyzed' AND decision = 'pending_review'"]
    else:
        conditions = ["status IN ('analyzed', 'pending')"]
        if recent_only:
            conditions.append("crawled_at >= NOW() - INTERVAL '10 days'")
        if decision:
            conditions.append("decision = %s")
            params.append(decision)
        if applied == "pending":
            conditions.append("user_status IS NULL")
        elif applied == "interested":
            conditions.append("user_status = 'interested'")
        elif applied == "applied":
            conditions.append("user_status = 'applied'")
        elif applied == "in_process":
            conditions.append("user_status = 'in_process'")
        elif applied == "offer":
            conditions.append("user_status = 'offer'")
        elif applied == "rejected":
            conditions.append("user_status = 'rejected'")
        elif applied == "dismissed":
            conditions.append("user_status = 'dismissed'")
        conditions.append("COALESCE(fit_score, 0) >= %s")
        params.append(min_score)
        conditions.append("COALESCE(fit_score, 0) <= %s")
        params.append(max_score)
    if search:
        conditions.append("(title ILIKE %s OR company ILIKE %s)")
        params.extend([f"%{search}%", f"%{search}%"])
    if hide_staffing:
        conditions.append("staffing_agency = FALSE")
    if hide_geo:
        conditions.append("geo_mismatch = FALSE")
    if english_only:
        conditions.append("likely_english = TRUE")
    if hide_rejected:
        conditions.append("(user_status IS NULL OR user_status NOT IN ('rejected', 'dismissed'))")
    return " AND ".join(conditions), params


@router.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return render(request, "jobs.html", {
        "stats": stats_service.get_stats(),
        **_jobs_filter_context(request),
    })


@router.get("/jobs", response_class=HTMLResponse)
async def list_jobs(
    request: Request,
    decision: str = Query("", alias="decision"),
    applied: str = Query("", alias="applied"),
    min_score: int = Query(0, alias="min_score"),
    max_score: int = Query(10, alias="max_score"),
    search: str = Query("", alias="search"),
    view: str = Query("", alias="view"),
    hide_staffing: bool = Query(False, alias="hide_staffing"),
    hide_geo: bool = Query(False, alias="hide_geo"),
    english_only: bool = Query(False, alias="english_only"),
    hide_rejected: bool = Query(True, alias="hide_rejected"),
    recent_only: bool = Query(True, alias="recent_only"),
    sort: str = Query("fit_score", alias="sort"),
    order: str = Query("desc", alias="order"),
    limit: int = Query(100, alias="limit"),
    offset: int = Query(0, alias="offset"),
):
    allowed_sorts = {
        "crawled_at", "fit_score", "company", "title", "location",
        "decision", "user_status", "analyzed_at",
    }
    sort_col = sort if sort in allowed_sorts else "fit_score"
    sort_dir = "ASC" if order.lower() == "asc" else "DESC"

    where, params = build_job_filter(
        decision, applied, min_score, max_score, search, view,
        hide_staffing=hide_staffing, hide_geo=hide_geo, english_only=english_only,
        hide_rejected=hide_rejected, recent_only=recent_only,
    )
    order_clause = build_order_clause(sort_col, sort_dir, view)
    query = (
        f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE {where} "
        f"ORDER BY {order_clause} LIMIT %s OFFSET %s"
    )
    params.extend([limit, offset])
    jobs = fetch_all(query, tuple(params))

    count_q = f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE {where}"
    total = fetch_one(count_q, tuple(params[:-2]))["cnt"]

    return render(request, "partials/job_rows.html", {
        "jobs": jobs, "total": total, "limit": limit, "offset": offset,
    })


@router.get("/export")
async def export_excel(
    decision: str = Query(""),
    applied: str = Query(""),
    min_score: int = Query(0),
    max_score: int = Query(10),
    search: str = Query(""),
    view: str = Query(""),
    hide_staffing: bool = Query(False),
    hide_geo: bool = Query(False),
    english_only: bool = Query(False),
    hide_rejected: bool = Query(True),
    recent_only: bool = Query(True),
    sort: str = Query("fit_score"),
    order: str = Query("desc"),
):
    allowed_sorts = {
        "crawled_at", "fit_score", "company", "title", "location",
        "decision", "user_status", "analyzed_at",
    }
    sort_col = sort if sort in allowed_sorts else "fit_score"
    sort_dir = "ASC" if order.lower() == "asc" else "DESC"

    where, params = build_job_filter(
        decision, applied, min_score, max_score, search, view,
        hide_staffing=hide_staffing, hide_geo=hide_geo, english_only=english_only,
        hide_rejected=hide_rejected, recent_only=recent_only,
    )
    order_clause = build_order_clause(sort_col, sort_dir, view)
    query = (
        f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE {where} "
        f"ORDER BY {order_clause}"
    )
    jobs = fetch_all(query, tuple(params))

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "Date", "Fetched At", "Posted", "Source", "Company", "Role", "Location", "Salary",
        "Score", "Decision", "Seniority", "Employment", "Start Date",
        "Blockers", "Strong Matches", "Reasoning", "Status", "Job URL",
        "Notes", "My Notes", "_dbId",
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    fill_active = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_interested = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    fill_action = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    fill_gray = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for row_idx, j in enumerate(jobs, 2):
        score = j["fit_score"] or 0
        us = j["user_status"] or ""
        blockers = "; ".join(j["hard_blockers"]) if j.get("hard_blockers") else ""
        matches = "; ".join(j["strong_matches"]) if j.get("strong_matches") else ""
        salary = format_salary(j)

        crawled = j.get("crawled_at")
        posted = j.get("posted_at")
        values = [
            crawled.strftime("%Y-%m-%d") if crawled else "",
            crawled.strftime("%Y-%m-%d %H:%M:%S UTC") if crawled else "",
            posted.strftime("%Y-%m-%d") if posted else "",
            j.get("source", ""),
            j.get("company", ""),
            j.get("title", ""),
            j.get("location", ""),
            salary,
            score,
            j.get("decision", ""),
            j.get("seniority_level", "") or "",
            j.get("employment_type", "") or "",
            j.get("start_date", "") or "",
            blockers,
            matches,
            j.get("reasoning", ""),
            us,
            j.get("url", ""),
            j.get("priority_notes", "") or "",
            j.get("notes", "") or "",
            j.get("id"),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        if us in ("applied", "in_process", "offer"):
            fill = fill_active
        elif us == "interested":
            fill = fill_interested
        elif us in ("rejected", "dismissed"):
            fill = fill_gray
        elif score >= 5 and not us:
            fill = fill_action
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14]:
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"jobs_export_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/jobs/new", response_class=HTMLResponse)
async def new_job_form(request: Request):
    return render(request, "new_job.html", {"form": {}})


@router.post("/jobs/new")
async def create_manual_job(
    request: Request,
    url: str = Form(...),
    title: str = Form(...),
    company: str = Form(...),
    location: str = Form(...),
    description: str = Form(...),
    remote: str = Form(""),
    salary_min: str = Form(""),
    salary_max: str = Form(""),
    salary_currency: str = Form("EUR"),
    tags: str = Form(""),
    already_applied: str = Form(""),
    applied_date: str = Form(""),
):
    url = url.strip()
    title = title.strip()
    company = company.strip()
    location = location.strip()
    description = description.strip()

    if not (url and title and company and location and description):
        form_values = {
            "url": url, "title": title, "company": company,
            "location": location, "description": description,
            "remote": bool(remote), "salary_min": salary_min,
            "salary_max": salary_max, "salary_currency": salary_currency,
            "tags": tags,
            "already_applied": bool(already_applied),
            "applied_date": applied_date,
        }
        return render(request, "new_job.html", {
            "form": form_values,
            "error": "All required fields must be filled.",
        })

    existing = fetch_one(f"SELECT id FROM {TABLE} WHERE url = %s", (url,))
    if existing:
        return RedirectResponse(
            url=f"/?duplicate={existing['id']}",
            status_code=303,
        )

    is_remote = bool(remote)
    s_min = int(salary_min) if salary_min.strip().isdigit() else None
    s_max = int(salary_max) if salary_max.strip().isdigit() else None
    tags_list = [t.strip() for t in tags.split(",") if t.strip()]
    tags_json = json.dumps(tags_list)

    execute(
        f"INSERT INTO {TABLE} "
        f"(url, title, company, location, description, remote, source, "
        f"external_id, description_quality, status, tags, likely_english, "
        f"salary_min, salary_max, salary_currency, posted_at) "
        f"VALUES (%s, %s, %s, %s, %s, %s, 'manual', NULL, 'good', 'pending', "
        f"%s::jsonb, FALSE, %s, %s, %s, NOW())",
        (url, title, company, location, description, is_remote,
         tags_json, s_min, s_max, salary_currency or "EUR"),
    )

    new_row = fetch_one(f"SELECT id FROM {TABLE} WHERE url = %s", (url,))
    new_id = new_row["id"] if new_row else 0

    if already_applied and new_id:
        applied_iso = (
            f"{applied_date}T12:00:00+00:00"
            if applied_date else
            datetime.now().astimezone().isoformat()
        )
        execute(
            f"UPDATE {TABLE} SET tracked_at = NOW() WHERE id = %s", (new_id,))
        execute(
            f"INSERT INTO {EVENTS_TABLE} (job_id, occurred_at, kind, label) "
            f"VALUES (%s, %s, 'application', 'Applied')",
            (new_id, applied_iso),
        )
        return RedirectResponse(url=f"/tracking#job-{new_id}", status_code=303)

    return RedirectResponse(url=f"/?new={new_id}", status_code=303)


@router.get("/jobs/{job_id}/view", response_class=HTMLResponse)
async def job_view(request: Request, job_id: int):
    job = fetch_one(f"SELECT * FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        return HTMLResponse("Job not found", status_code=404)
    return render(request, "job_view.html", {"job": job})


@router.get("/jobs/{job_id}", response_class=HTMLResponse)
async def job_detail(request: Request, job_id: int):
    job = fetch_one(f"SELECT * FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        return HTMLResponse("<tr><td colspan='9'>Job not found</td></tr>", status_code=404)
    return render(request, "partials/job_detail.html", {"job": job})


@router.patch("/jobs/{job_id}", response_class=HTMLResponse)
async def update_job(request: Request, job_id: int, decision: str = Form(...)):
    if decision not in ("PASS", "SKIP", "MAYBE"):
        return HTMLResponse("Invalid decision", status_code=400)
    execute(
        f"UPDATE {TABLE} SET decision = %s WHERE id = %s",
        (decision, job_id),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.patch("/jobs/{job_id}/status", response_class=HTMLResponse)
async def set_user_status(request: Request, job_id: int, user_status: str = Form(...)):
    valid = ("interested", "applied", "in_process", "offer", "rejected", "dismissed", "")
    if user_status not in valid:
        return HTMLResponse("Invalid status", status_code=400)
    status_val = user_status if user_status else None
    execute(
        f"UPDATE {TABLE} SET user_status = %s, "
        f"applied_at = CASE WHEN %s = 'applied' THEN NOW() ELSE applied_at END, "
        f"sheet_synced = FALSE WHERE id = %s",
        (status_val, status_val, job_id),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.patch("/jobs/{job_id}/notes", response_class=HTMLResponse)
async def update_notes(request: Request, job_id: int, notes: str = Form("")):
    execute(
        f"UPDATE {TABLE} SET notes = %s, sheet_synced = FALSE WHERE id = %s",
        (notes if notes.strip() else None, job_id),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.patch("/jobs/{job_id}/review", response_class=HTMLResponse)
async def resolve_review(
    request: Request,
    job_id: int,
    review_action: str = Form(...),
    notes: str = Form(""),
):
    action = REVIEW_ACTIONS.get(review_action)
    if not action:
        return HTMLResponse("Invalid review action", status_code=400)

    review_flag_clause = "needs_human_review = FALSE, " if schema.HAS_HUMAN_REVIEW_COLUMNS else ""
    event_notes = (notes or "").strip() or None

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"SELECT id, notes FROM {TABLE} WHERE id = %s",
                (job_id,),
            )
            existing = cur.fetchone()
            if not existing:
                return HTMLResponse("Job not found", status_code=404)

            next_notes = _review_notes_value(existing.get("notes"), notes)
            cur.execute(
                f"""
                UPDATE {TABLE}
                SET decision = %s,
                    user_status = %s,
                    {review_flag_clause}notes = %s,
                    sheet_synced = FALSE
                WHERE id = %s
                """,
                (action["decision"], action["user_status"], next_notes, job_id),
            )
            cur.execute(
                f"""
                INSERT INTO {EVENTS_TABLE} (job_id, occurred_at, kind, label, notes)
                VALUES (%s, NOW(), 'decision', %s, %s)
                """,
                (job_id, action["label"], event_notes),
            )
            cur.execute(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
            job = cur.fetchone()

    return render(request, "partials/job_row_single.html", {"job": job})


@router.post("/jobs/{job_id}/rescore", response_class=HTMLResponse)
async def rescore_job(request: Request, job_id: int, description: str = Form(...)):
    execute(
        f"UPDATE {TABLE} SET description = %s, status = 'pending', "
        f"analyzed_at = NULL, error = NULL, retry_count = 0, "
        f"sheet_synced = FALSE, sheet_synced_at = NULL WHERE id = %s",
        (description, job_id),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.post("/jobs/{job_id}/retry", response_class=HTMLResponse)
async def retry_job(request: Request, job_id: int):
    execute(
        f"UPDATE {TABLE} SET status = 'pending', retry_count = 0, "
        f"error = NULL, error_code = NULL WHERE id = %s",
        (job_id,),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.patch("/jobs/{job_id}/duplicate", response_class=HTMLResponse)
async def confirm_duplicate(request: Request, job_id: int, confirmed: str = Form(...)):
    if confirmed not in ("true", "false"):
        return HTMLResponse("Invalid value", status_code=400)
    execute(
        f"UPDATE {TABLE} SET duplicate_confirmed = %s WHERE id = %s",
        (confirmed == "true", job_id),
    )
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@router.post("/jobs/{job_id}/track", response_class=HTMLResponse)
async def htmx_start_tracking(request: Request, job_id: int):
    """HTMX target: start tracking and re-render the row.

    Mirrors POST /api/tracking/jobs/{id}/start but returns the row partial.
    """
    job = fetch_one(f"SELECT id, tracked_at FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        return HTMLResponse("Not found", status_code=404)
    if job["tracked_at"] is None:
        execute(f"UPDATE {TABLE} SET tracked_at = NOW() WHERE id = %s", (job_id,))
    job = fetch_one(f"SELECT {schema.ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})
