"""Jobs Funnel — lightweight review UI.

Start:
    cd D:/projects/jobs_funnel
    python -m uvicorn ui.server:app --port 8080 --reload
"""

import io
import json
from datetime import datetime
from urllib.parse import urlencode

import psycopg2
import psycopg2.extras
from fastapi import FastAPI, Form, HTTPException, Query, Request
from pydantic import BaseModel
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from scripts import calibration_proposals
from scripts import calibration_settings
from ui.config import EVENTS_TABLE, STATIC_DIR, TABLE, TEMPLATES_DIR
from ui.db import execute, fetch_all, fetch_one, get_db
from ui.rendering import format_salary, render, templates
from ui.services.calibration_presenter import proposal_summary_lines
from ui.services import system_health

# ── Config ───────────────────────────────────────────────────────────
app = FastAPI(title="Jobs Funnel UI")
STATIC_DIR.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# ── Optional schema detection ───────────────────────────────────────
# The Phase 1 embedding migration (0003_pgvector.sql) is per-profile-table.
# Human review columns are also optional while Phase 4 migrations roll out.
def _detect_optional_columns():
    wanted = {
        "embedding",
        "scored_uncalibrated",
        "needs_human_review",
        "explanation",
        "confidence",
        "critique_count",
    }
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = %s AND column_name = ANY(%s)",
                    (TABLE, list(wanted)),
                )
                return {r[0] for r in cur.fetchall()}
    except HTTPException:
        return set()


OPTIONAL_COLUMNS = _detect_optional_columns()
HAS_EMBEDDING_COLUMNS = {"embedding", "scored_uncalibrated"}.issubset(OPTIONAL_COLUMNS)
HAS_HUMAN_REVIEW_COLUMNS = {
    "needs_human_review",
    "explanation",
    "confidence",
    "critique_count",
}.issubset(OPTIONAL_COLUMNS)

REVIEW_ACTIONS = {
    "apply_target": {
        "decision": "PASS",
        "user_status": "interested",
        "label": "Reviewed: apply target",
    },
    "maybe": {
        "decision": "MAYBE",
        "user_status": "interested",
        "label": "Reviewed: maybe",
    },
    "skip": {
        "decision": "SKIP",
        "user_status": "dismissed",
        "label": "Reviewed: skip",
    },
}


def _review_notes_value(existing_notes, submitted_notes):
    submitted = (submitted_notes or "").strip()
    if submitted:
        return submitted
    return existing_notes if existing_notes else None

_BASE_ROW_COLS = (
    "id, url, title, company, location, source, fit_score, decision, "
    "cv_variant, reasoning, status, crawled_at, analyzed_at, "
    "salary_min, salary_max, salary_currency, remote, likely_english, "
    "staffing_agency, geo_mismatch, "
    "tags, priority_notes, notes, user_status, "
    "posted_at, employment_type, seniority_level, start_date, "
    "error, error_code, retry_count, "
    "possible_duplicate_of, duplicate_confirmed, "
    "tracked_at"
)

if HAS_EMBEDDING_COLUMNS:
    ROW_COLS = (
        f"{_BASE_ROW_COLS}, "
        "(embedding IS NULL) AS awaiting_embedding, scored_uncalibrated"
    )
else:
    ROW_COLS = (
        f"{_BASE_ROW_COLS}, "
        "FALSE AS awaiting_embedding, FALSE AS scored_uncalibrated"
    )

if HAS_HUMAN_REVIEW_COLUMNS:
    ROW_COLS = (
        f"{ROW_COLS}, needs_human_review, explanation, confidence, critique_count"
    )
else:
    ROW_COLS = (
        f"{ROW_COLS}, FALSE AS needs_human_review, NULL AS explanation, "
        "NULL AS confidence, 0 AS critique_count"
    )

def _query_bool(request: Request, name: str, default: bool = False) -> bool:
    value = request.query_params.get(name)
    if value is None:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def _jobs_filter_context(request: Request):
    filters = {
        "view": request.query_params.get("view", ""),
        "decision": request.query_params.get("decision", ""),
        "applied": request.query_params.get("applied", ""),
        "min_score": request.query_params.get("min_score", "0"),
        "max_score": request.query_params.get("max_score", "10"),
        "search": request.query_params.get("search", ""),
        "hide_staffing": _query_bool(request, "hide_staffing"),
        "hide_geo": _query_bool(request, "hide_geo"),
        "english_only": _query_bool(request, "english_only"),
        "recent_only": _query_bool(request, "recent_only", True),
        "hide_rejected": _query_bool(request, "hide_rejected", True),
        "sort": request.query_params.get("sort", "fit_score"),
        "order": request.query_params.get("order", "desc"),
    }
    query = urlencode({
        key: str(value).lower() if isinstance(value, bool) else value
        for key, value in filters.items()
        if value not in ("", None)
    })
    return {"filters": filters, "jobs_query_string": query}


def _calibration_context(error: str | None = None):
    active = calibration_settings.load_active_settings(force=True)
    proposals = calibration_proposals.list_proposals(limit=20)
    proposals = [
        {**proposal, "summary_lines": proposal_summary_lines(proposal)}
        for proposal in proposals
    ]
    return {
        "active": active,
        "proposals": proposals,
        "error": error,
    }


def _render_calibration_content(request: Request, error: str | None = None):
    return render(request, "partials/calibration_content.html", _calibration_context(error))

# ── Stats helper ─────────────────────────────────────────────────────
def get_stats():
    rows = fetch_all(
        f"SELECT decision, COUNT(*) as cnt FROM {TABLE} "
        f"WHERE status = 'analyzed' GROUP BY decision"
    )
    stats = {"total": 0, "PASS": 0, "MAYBE": 0, "SKIP": 0}
    for r in rows:
        stats[r["decision"]] = r["cnt"]
        stats["total"] += r["cnt"]
    pending = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE status = 'pending'"
    )
    stats["pending"] = pending["cnt"] if pending else 0
    interested = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE user_status = 'interested'"
    )
    stats["interested"] = interested["cnt"] if interested else 0
    applied = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE user_status IN ('applied', 'in_process', 'offer')"
    )
    stats["applied"] = applied["cnt"] if applied else 0
    dismissed = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE user_status IN ('dismissed', 'rejected')"
    )
    stats["dismissed"] = dismissed["cnt"] if dismissed else 0
    error = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE status = 'error'"
    )
    stats["error"] = error["cnt"] if error else 0
    dead = fetch_one(
        f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE status = 'dead'"
    )
    stats["dead"] = dead["cnt"] if dead else 0
    if HAS_EMBEDDING_COLUMNS:
        awaiting = fetch_one(
            f"SELECT COUNT(*) as cnt FROM {TABLE} "
            f"WHERE embedding IS NULL AND (error_code IS NULL OR error_code != 'EMBED_FAILED')"
        )
        stats["awaiting_embedding"] = awaiting["cnt"] if awaiting else 0
    else:
        stats["awaiting_embedding"] = 0
    return stats


# ── Query builder ────────────────────────────────────────────────────
def build_order_clause(sort_col: str, sort_dir: str, view: str = "") -> str:
    if view == "review":
        return "COALESCE(fit_score, 0) DESC, analyzed_at DESC NULLS LAST, id DESC"
    return f"{sort_col} {sort_dir}, id {sort_dir}"


def build_job_filter(decision="", applied="", min_score=0, max_score=10, search="", view="",
                     hide_staffing=False, hide_geo=False, english_only=False,
                     hide_rejected=False, recent_only=True):
    params: list = []
    if view == "error":
        conditions = ["status = 'error'"]
    elif view == "dead":
        conditions = ["status = 'dead'"]
    elif view == "failed":
        conditions = ["status IN ('error', 'dead')"]
    elif view == "duplicates":
        conditions = ["possible_duplicate_of IS NOT NULL AND duplicate_confirmed IS NULL"]
    elif view == "review":
        if HAS_HUMAN_REVIEW_COLUMNS:
            conditions = ["status = 'analyzed' AND (needs_human_review = TRUE OR decision = 'pending_review')"]
        else:
            conditions = ["status = 'analyzed' AND decision = 'pending_review'"]
    else:
        conditions = ["status IN ('analyzed', 'pending')"]
        if recent_only:
            conditions.append("crawled_at >= NOW() - INTERVAL '10 days'")
        if decision:
            conditions.append("decision = %s")
            params.append(decision)
        if applied == "pending":
            conditions.append("user_status IS NULL")
        elif applied == "interested":
            conditions.append("user_status = 'interested'")
        elif applied == "applied":
            conditions.append("user_status = 'applied'")
        elif applied == "in_process":
            conditions.append("user_status = 'in_process'")
        elif applied == "offer":
            conditions.append("user_status = 'offer'")
        elif applied == "rejected":
            conditions.append("user_status = 'rejected'")
        elif applied == "dismissed":
            conditions.append("user_status = 'dismissed'")
        conditions.append("COALESCE(fit_score, 0) >= %s")
        params.append(min_score)
        conditions.append("COALESCE(fit_score, 0) <= %s")
        params.append(max_score)
    if search:
        conditions.append("(title ILIKE %s OR company ILIKE %s)")
        params.extend([f"%{search}%", f"%{search}%"])
    if hide_staffing:
        conditions.append("staffing_agency = FALSE")
    if hide_geo:
        conditions.append("geo_mismatch = FALSE")
    if english_only:
        conditions.append("likely_english = TRUE")
    if hide_rejected:
        conditions.append("(user_status IS NULL OR user_status NOT IN ('rejected', 'dismissed'))")
    return " AND ".join(conditions), params


# ── Routes ───────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return render(request, "jobs.html", {
        "stats": get_stats(),
        **_jobs_filter_context(request),
    })


@app.get("/jobs", response_class=HTMLResponse)
async def list_jobs(
    request: Request,
    decision: str = Query("", alias="decision"),
    applied: str = Query("", alias="applied"),
    min_score: int = Query(0, alias="min_score"),
    max_score: int = Query(10, alias="max_score"),
    search: str = Query("", alias="search"),
    view: str = Query("", alias="view"),
    hide_staffing: bool = Query(False, alias="hide_staffing"),
    hide_geo: bool = Query(False, alias="hide_geo"),
    english_only: bool = Query(False, alias="english_only"),
    hide_rejected: bool = Query(True, alias="hide_rejected"),
    recent_only: bool = Query(True, alias="recent_only"),
    sort: str = Query("fit_score", alias="sort"),
    order: str = Query("desc", alias="order"),
    limit: int = Query(100, alias="limit"),
    offset: int = Query(0, alias="offset"),
):
    allowed_sorts = {
        "crawled_at", "fit_score", "company", "title", "location",
        "decision", "user_status", "analyzed_at",
    }
    sort_col = sort if sort in allowed_sorts else "fit_score"
    sort_dir = "ASC" if order.lower() == "asc" else "DESC"

    where, params = build_job_filter(
        decision, applied, min_score, max_score, search, view,
        hide_staffing=hide_staffing, hide_geo=hide_geo, english_only=english_only,
        hide_rejected=hide_rejected, recent_only=recent_only,
    )
    # Add id as tiebreaker so pagination is stable (avoids duplicates on Load More)
    order_clause = build_order_clause(sort_col, sort_dir, view)
    query = (
        f"SELECT {ROW_COLS} FROM {TABLE} WHERE {where} "
        f"ORDER BY {order_clause} LIMIT %s OFFSET %s"
    )
    params.extend([limit, offset])
    jobs = fetch_all(query, tuple(params))

    count_q = f"SELECT COUNT(*) as cnt FROM {TABLE} WHERE {where}"
    total = fetch_one(count_q, tuple(params[:-2]))["cnt"]

    return render(request, "partials/job_rows.html", {
        "jobs": jobs, "total": total, "limit": limit, "offset": offset,
    })


@app.get("/export")
async def export_excel(
    decision: str = Query(""),
    applied: str = Query(""),
    min_score: int = Query(0),
    max_score: int = Query(10),
    search: str = Query(""),
    view: str = Query(""),
    hide_staffing: bool = Query(False),
    hide_geo: bool = Query(False),
    english_only: bool = Query(False),
    hide_rejected: bool = Query(True),
    recent_only: bool = Query(True),
    sort: str = Query("fit_score"),
    order: str = Query("desc"),
):
    allowed_sorts = {
        "crawled_at", "fit_score", "company", "title", "location",
        "decision", "user_status", "analyzed_at",
    }
    sort_col = sort if sort in allowed_sorts else "fit_score"
    sort_dir = "ASC" if order.lower() == "asc" else "DESC"

    where, params = build_job_filter(
        decision, applied, min_score, max_score, search, view,
        hide_staffing=hide_staffing, hide_geo=hide_geo, english_only=english_only,
        hide_rejected=hide_rejected, recent_only=recent_only,
    )
    order_clause = build_order_clause(sort_col, sort_dir, view)
    query = (
        f"SELECT {ROW_COLS} FROM {TABLE} WHERE {where} "
        f"ORDER BY {order_clause}"
    )
    jobs = fetch_all(query, tuple(params))

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "Date", "Fetched At", "Posted", "Source", "Company", "Role", "Location", "Salary",
        "Score", "Decision", "Seniority", "Employment", "Start Date",
        "Blockers", "Strong Matches", "Reasoning", "Status", "Job URL",
        "Notes", "My Notes", "_dbId",
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    fill_active = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_interested = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    fill_action = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    fill_gray = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for row_idx, j in enumerate(jobs, 2):
        score = j["fit_score"] or 0
        us = j["user_status"] or ""
        blockers = "; ".join(j["hard_blockers"]) if j.get("hard_blockers") else ""
        matches = "; ".join(j["strong_matches"]) if j.get("strong_matches") else ""
        salary = format_salary(j)

        crawled = j.get("crawled_at")
        posted = j.get("posted_at")
        values = [
            crawled.strftime("%Y-%m-%d") if crawled else "",
            crawled.strftime("%Y-%m-%d %H:%M:%S UTC") if crawled else "",
            posted.strftime("%Y-%m-%d") if posted else "",
            j.get("source", ""),
            j.get("company", ""),
            j.get("title", ""),
            j.get("location", ""),
            salary,
            score,
            j.get("decision", ""),
            j.get("seniority_level", "") or "",
            j.get("employment_type", "") or "",
            j.get("start_date", "") or "",
            blockers,
            matches,
            j.get("reasoning", ""),
            us,
            j.get("url", ""),
            j.get("priority_notes", "") or "",
            j.get("notes", "") or "",
            j.get("id"),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        # Row coloring
        if us in ("applied", "in_process", "offer"):
            fill = fill_active
        elif us == "interested":
            fill = fill_interested
        elif us in ("rejected", "dismissed"):
            fill = fill_gray
        elif score >= 5 and not us:
            fill = fill_action
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # Auto-width for key columns
    for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14]:
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"jobs_export_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/jobs/new", response_class=HTMLResponse)
async def new_job_form(request: Request):
    return render(request, "new_job.html", {"form": {}})


@app.post("/jobs/new")
async def create_manual_job(
    request: Request,
    url: str = Form(...),
    title: str = Form(...),
    company: str = Form(...),
    location: str = Form(...),
    description: str = Form(...),
    remote: str = Form(""),
    salary_min: str = Form(""),
    salary_max: str = Form(""),
    salary_currency: str = Form("EUR"),
    tags: str = Form(""),
    already_applied: str = Form(""),
    applied_date: str = Form(""),
):
    url = url.strip()
    title = title.strip()
    company = company.strip()
    location = location.strip()
    description = description.strip()

    if not (url and title and company and location and description):
        form_values = {
            "url": url, "title": title, "company": company,
            "location": location, "description": description,
            "remote": bool(remote), "salary_min": salary_min,
            "salary_max": salary_max, "salary_currency": salary_currency,
            "tags": tags,
            "already_applied": bool(already_applied),
            "applied_date": applied_date,
        }
        return render(request, "new_job.html", {
            "form": form_values,
            "error": "All required fields must be filled.",
        })

    existing = fetch_one(f"SELECT id FROM {TABLE} WHERE url = %s", (url,))
    if existing:
        return RedirectResponse(
            url=f"/?duplicate={existing['id']}",
            status_code=303,
        )

    is_remote = bool(remote)
    s_min = int(salary_min) if salary_min.strip().isdigit() else None
    s_max = int(salary_max) if salary_max.strip().isdigit() else None
    tags_list = [t.strip() for t in tags.split(",") if t.strip()]
    tags_json = json.dumps(tags_list)

    execute(
        f"INSERT INTO {TABLE} "
        f"(url, title, company, location, description, remote, source, "
        f"external_id, description_quality, status, tags, likely_english, "
        f"salary_min, salary_max, salary_currency, posted_at) "
        f"VALUES (%s, %s, %s, %s, %s, %s, 'manual', NULL, 'good', 'pending', "
        f"%s::jsonb, FALSE, %s, %s, %s, NOW())",
        (url, title, company, location, description, is_remote,
         tags_json, s_min, s_max, salary_currency or "EUR"),
    )

    new_row = fetch_one(f"SELECT id FROM {TABLE} WHERE url = %s", (url,))
    new_id = new_row["id"] if new_row else 0

    if already_applied and new_id:
        applied_iso = (
            f"{applied_date}T12:00:00+00:00"
            if applied_date else
            datetime.now().astimezone().isoformat()
        )
        execute(
            f"UPDATE {TABLE} SET tracked_at = NOW() WHERE id = %s", (new_id,))
        execute(
            f"INSERT INTO {EVENTS_TABLE} (job_id, occurred_at, kind, label) "
            f"VALUES (%s, %s, 'application', 'Applied')",
            (new_id, applied_iso),
        )
        return RedirectResponse(url=f"/tracking#job-{new_id}", status_code=303)

    return RedirectResponse(url=f"/?new={new_id}", status_code=303)


@app.get("/jobs/{job_id}", response_class=HTMLResponse)
async def job_detail(request: Request, job_id: int):
    job = fetch_one(f"SELECT * FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        return HTMLResponse("<tr><td colspan='9'>Job not found</td></tr>", status_code=404)
    return render(request, "partials/job_detail.html", {"job": job})


@app.patch("/jobs/{job_id}", response_class=HTMLResponse)
async def update_job(request: Request, job_id: int, decision: str = Form(...)):
    if decision not in ("PASS", "SKIP", "MAYBE"):
        return HTMLResponse("Invalid decision", status_code=400)
    execute(
        f"UPDATE {TABLE} SET decision = %s WHERE id = %s",
        (decision, job_id),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.patch("/jobs/{job_id}/status", response_class=HTMLResponse)
async def set_user_status(request: Request, job_id: int, user_status: str = Form(...)):
    valid = ("interested", "applied", "in_process", "offer", "rejected", "dismissed", "")
    if user_status not in valid:
        return HTMLResponse("Invalid status", status_code=400)
    status_val = user_status if user_status else None
    execute(
        f"UPDATE {TABLE} SET user_status = %s, "
        f"applied_at = CASE WHEN %s = 'applied' THEN NOW() ELSE applied_at END, "
        f"sheet_synced = FALSE WHERE id = %s",
        (status_val, status_val, job_id),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.patch("/jobs/{job_id}/notes", response_class=HTMLResponse)
async def update_notes(request: Request, job_id: int, notes: str = Form("")):
    execute(
        f"UPDATE {TABLE} SET notes = %s, sheet_synced = FALSE WHERE id = %s",
        (notes if notes.strip() else None, job_id),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.patch("/jobs/{job_id}/review", response_class=HTMLResponse)
async def resolve_review(
    request: Request,
    job_id: int,
    review_action: str = Form(...),
    notes: str = Form(""),
):
    action = REVIEW_ACTIONS.get(review_action)
    if not action:
        return HTMLResponse("Invalid review action", status_code=400)

    review_flag_clause = "needs_human_review = FALSE, " if HAS_HUMAN_REVIEW_COLUMNS else ""
    event_notes = (notes or "").strip() or None

    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"SELECT id, notes FROM {TABLE} WHERE id = %s",
                (job_id,),
            )
            existing = cur.fetchone()
            if not existing:
                return HTMLResponse("Job not found", status_code=404)

            next_notes = _review_notes_value(existing.get("notes"), notes)
            cur.execute(
                f"""
                UPDATE {TABLE}
                SET decision = %s,
                    user_status = %s,
                    {review_flag_clause}notes = %s,
                    sheet_synced = FALSE
                WHERE id = %s
                """,
                (action["decision"], action["user_status"], next_notes, job_id),
            )
            cur.execute(
                f"""
                INSERT INTO {EVENTS_TABLE} (job_id, occurred_at, kind, label, notes)
                VALUES (%s, NOW(), 'decision', %s, %s)
                """,
                (job_id, action["label"], event_notes),
            )
            cur.execute(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
            job = cur.fetchone()

    return render(request, "partials/job_row_single.html", {"job": job})


@app.post("/jobs/{job_id}/rescore", response_class=HTMLResponse)
async def rescore_job(request: Request, job_id: int, description: str = Form(...)):
    execute(
        f"UPDATE {TABLE} SET description = %s, status = 'pending', "
        f"analyzed_at = NULL, error = NULL, retry_count = 0, "
        f"sheet_synced = FALSE, sheet_synced_at = NULL WHERE id = %s",
        (description, job_id),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.post("/jobs/{job_id}/retry", response_class=HTMLResponse)
async def retry_job(request: Request, job_id: int):
    execute(
        f"UPDATE {TABLE} SET status = 'pending', retry_count = 0, "
        f"error = NULL, error_code = NULL WHERE id = %s",
        (job_id,),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.patch("/jobs/{job_id}/duplicate", response_class=HTMLResponse)
async def confirm_duplicate(request: Request, job_id: int, confirmed: str = Form(...)):
    if confirmed not in ("true", "false"):
        return HTMLResponse("Invalid value", status_code=400)
    execute(
        f"UPDATE {TABLE} SET duplicate_confirmed = %s WHERE id = %s",
        (confirmed == "true", job_id),
    )
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.post("/jobs/{job_id}/track", response_class=HTMLResponse)
async def htmx_start_tracking(request: Request, job_id: int):
    """HTMX target: start tracking and re-render the row.

    Mirrors POST /api/tracking/jobs/{id}/start but returns the row partial.
    """
    job = fetch_one(f"SELECT id, tracked_at FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        return HTMLResponse("Not found", status_code=404)
    if job["tracked_at"] is None:
        execute(f"UPDATE {TABLE} SET tracked_at = NOW() WHERE id = %s", (job_id,))
    job = fetch_one(f"SELECT {ROW_COLS} FROM {TABLE} WHERE id = %s", (job_id,))
    return render(request, "partials/job_row_single.html", {"job": job})


@app.get("/stats", response_class=HTMLResponse)
async def stats(request: Request):
    return render(request, "partials/stats.html", {"stats": get_stats()})


# ── Pipeline Runs ────────────────────────────────────────────────────
@app.get("/runs", response_class=HTMLResponse)
async def runs_page(request: Request):
    return render(request, "runs.html")


@app.get("/calibration", response_class=HTMLResponse)
async def calibration_page(request: Request):
    return render(request, "calibration.html", _calibration_context())


@app.get("/system", response_class=HTMLResponse)
def system_page(request: Request):
    return render(request, "system.html", {"checks": system_health.collect_system_health()})


@app.post("/calibration/proposals", response_class=HTMLResponse)
async def calibration_generate_proposal(
    request: Request,
    window_days: int = Form(90),
):
    try:
        calibration_proposals.generate_proposal(window_days=window_days)
    except calibration_proposals.ProposalStateError as exc:
        return _render_calibration_content(request, str(exc))
    return _render_calibration_content(request)


@app.post("/calibration/proposals/{proposal_id}/apply", response_class=HTMLResponse)
async def calibration_apply_proposal(request: Request, proposal_id: int):
    try:
        calibration_proposals.apply_proposal(proposal_id)
    except calibration_proposals.ProposalStateError as exc:
        return _render_calibration_content(request, str(exc))
    return _render_calibration_content(request)


@app.post("/calibration/proposals/{proposal_id}/rollback", response_class=HTMLResponse)
async def calibration_rollback_proposal(request: Request, proposal_id: int):
    try:
        calibration_proposals.rollback_proposal(proposal_id)
    except calibration_proposals.ProposalStateError as exc:
        return _render_calibration_content(request, str(exc))
    return _render_calibration_content(request)


@app.get("/tracking", response_class=HTMLResponse)
async def tracking_page(request: Request):
    return render(request, "tracking.html")


@app.get("/runs/list", response_class=HTMLResponse)
async def runs_list(
    request: Request,
    limit: int = Query(50, alias="limit"),
    offset: int = Query(0, alias="offset"),
):
    runs = fetch_all(
        "SELECT * FROM pipeline_runs ORDER BY started_at DESC LIMIT %s OFFSET %s",
        (limit, offset),
    )
    total = fetch_one("SELECT COUNT(*) as cnt FROM pipeline_runs")["cnt"]
    return render(request, "partials/run_rows.html", {
        "runs": runs, "total": total, "limit": limit, "offset": offset,
    })


# ── Tracking API ─────────────────────────────────────────────────────
def _serialize_job_with_events(row):
    """Shape a joined jobs+events row dict into the API response format."""
    return {
        "id": row["id"],
        "title": row["title"],
        "company": row["company"],
        "url": row["url"],
        "location": row["location"],
        "tracked_at": row["tracked_at"].isoformat() if row["tracked_at"] else None,
        "closed_at": row["closed_at"].isoformat() if row.get("closed_at") else None,
        "user_status": row["user_status"],
        "events": [],
    }


@app.get("/api/tracking/jobs")
async def api_tracking_jobs():
    """Return all tracked jobs with their events embedded.

    Ordered so the job whose latest event is newest appears first.
    Jobs with no events yet sort by tracked_at desc.
    """
    jobs = fetch_all(
        f"SELECT id, title, company, url, location, tracked_at, closed_at, user_status "
        f"FROM {TABLE} WHERE tracked_at IS NOT NULL"
    )
    if not jobs:
        return []
    by_id = {j["id"]: _serialize_job_with_events(j) for j in jobs}

    events = fetch_all(
        f"SELECT id, job_id, occurred_at, kind, label, notes "
        f"FROM {EVENTS_TABLE} WHERE job_id = ANY(%s) "
        f"ORDER BY occurred_at ASC, id ASC",
        (list(by_id.keys()),),
    )
    for ev in events:
        if ev["job_id"] in by_id:
            by_id[ev["job_id"]]["events"].append({
                "id": ev["id"],
                "occurred_at": ev["occurred_at"].isoformat(),
                "kind": ev["kind"],
                "label": ev["label"],
                "notes": ev["notes"],
            })

    def latest_event_ts(job):
        if job["events"]:
            return job["events"][-1]["occurred_at"]
        return job["tracked_at"] or ""

    result = sorted(by_id.values(), key=latest_event_ts, reverse=True)
    return result


@app.post("/api/tracking/jobs/{job_id}/start")
async def api_tracking_start(job_id: int):
    job = fetch_one(f"SELECT id, tracked_at FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job["tracked_at"] is None:
        execute(f"UPDATE {TABLE} SET tracked_at = NOW() WHERE id = %s", (job_id,))
        job = fetch_one(f"SELECT tracked_at FROM {TABLE} WHERE id = %s", (job_id,))
    return {"tracked_at": job["tracked_at"].isoformat()}


@app.post("/api/tracking/jobs/{job_id}/stop")
async def api_tracking_stop(job_id: int):
    job = fetch_one(f"SELECT id FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    execute(f"UPDATE {TABLE} SET tracked_at = NULL WHERE id = %s", (job_id,))
    return {}


@app.post("/api/tracking/jobs/{job_id}/close")
async def api_tracking_close(job_id: int):
    job = fetch_one(f"SELECT id, closed_at FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job["closed_at"] is None:
        execute(f"UPDATE {TABLE} SET closed_at = NOW() WHERE id = %s", (job_id,))
        job = fetch_one(f"SELECT closed_at FROM {TABLE} WHERE id = %s", (job_id,))
    return {"closed_at": job["closed_at"].isoformat()}


@app.post("/api/tracking/jobs/{job_id}/reopen")
async def api_tracking_reopen(job_id: int):
    job = fetch_one(f"SELECT id FROM {TABLE} WHERE id = %s", (job_id,))
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    execute(f"UPDATE {TABLE} SET closed_at = NULL WHERE id = %s", (job_id,))
    return {}


class EventCreate(BaseModel):
    job_id: int
    occurred_at: str
    kind: str
    label: str
    notes: str | None = None


class EventUpdate(BaseModel):
    occurred_at: str | None = None
    kind: str | None = None
    label: str | None = None
    notes: str | None = None


VALID_EVENT_KINDS = {"application", "contact", "interview", "task", "decision", "note"}


def _serialize_event(row):
    return {
        "id": row["id"],
        "job_id": row["job_id"],
        "occurred_at": row["occurred_at"].isoformat(),
        "kind": row["kind"],
        "label": row["label"],
        "notes": row["notes"],
    }


@app.post("/api/tracking/events")
async def api_create_event(payload: EventCreate):
    if payload.kind not in VALID_EVENT_KINDS:
        raise HTTPException(status_code=400, detail=f"Invalid kind: {payload.kind}")
    if not payload.label.strip():
        raise HTTPException(status_code=400, detail="label is required")
    job = fetch_one(f"SELECT id FROM {TABLE} WHERE id = %s", (payload.job_id,))
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    row = fetch_one(
        f"INSERT INTO {EVENTS_TABLE} (job_id, occurred_at, kind, label, notes) "
        f"VALUES (%s, %s, %s, %s, %s) "
        f"RETURNING id, job_id, occurred_at, kind, label, notes",
        (payload.job_id, payload.occurred_at, payload.kind,
         payload.label.strip(), payload.notes),
    )
    return _serialize_event(row)


@app.patch("/api/tracking/events/{event_id}")
async def api_update_event(event_id: int, payload: EventUpdate):
    existing = fetch_one(
        f"SELECT id, job_id, occurred_at, kind, label, notes "
        f"FROM {EVENTS_TABLE} WHERE id = %s", (event_id,))
    if not existing:
        raise HTTPException(status_code=404, detail="Event not found")
    if payload.kind is not None and payload.kind not in VALID_EVENT_KINDS:
        raise HTTPException(status_code=400, detail=f"Invalid kind: {payload.kind}")

    fields = []
    values: list = []
    for col in ("occurred_at", "kind", "label", "notes"):
        v = getattr(payload, col)
        if v is not None:
            fields.append(f"{col} = %s")
            values.append(v)
    if not fields:
        return _serialize_event(existing)
    values.append(event_id)
    row = fetch_one(
        f"UPDATE {EVENTS_TABLE} SET {', '.join(fields)} WHERE id = %s "
        f"RETURNING id, job_id, occurred_at, kind, label, notes",
        tuple(values),
    )
    return _serialize_event(row)


@app.delete("/api/tracking/events/{event_id}")
async def api_delete_event(event_id: int):
    existing = fetch_one(f"SELECT id FROM {EVENTS_TABLE} WHERE id = %s", (event_id,))
    if not existing:
        raise HTTPException(status_code=404, detail="Event not found")
    execute(f"DELETE FROM {EVENTS_TABLE} WHERE id = %s", (event_id,))
    return {}
